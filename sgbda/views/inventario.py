import re
from django.shortcuts import render, redirect
from django.contrib import messages
from sgbda.models import instancia, cliente, servidor, checkSQL, actualizaciones
from sgbda.services.actualizar_instancias_v3 import actualizar_instancias_desde_conexiones
from django.utils import timezone
from django.db.models import Q
import queue
import threading
import re
from django.http import StreamingHttpResponse
# from django.views.decorators.http import require_GET
from django.http import HttpResponse
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl import Workbook
from openpyxl.utils import get_column_letter


#=================================================================================
#=================================================================================
#=================================================================================
#=================================================================================

def listar_inventario(request):

    buscar = request.GET.get('buscar', '')
    version = request.GET.get('versiones_sql', '')
    version_pg = request.GET.get('versiones_pg', '')
    clientes_obj = request.GET.get('cliente_actual', '')
    administrado = request.GET.get('administrados')

    all_clientes = cliente.objects.all()

    all_inventario =  instancia.objects.select_related(
        'servidor', 
        'servidor__cliente'
    ).prefetch_related(
        'servicios',
        'checksql'
    ).all()

    # FILTRO DE TEXTO
    if buscar:
        all_inventario = all_inventario.filter(
            Q(nombre_instancia__icontains=buscar) |
            Q(build__icontains=buscar) |
            Q(edition__icontains=buscar) |
            Q(servidor__ip__icontains=buscar) |
            Q(servidor__hostname__icontains=buscar) |
            Q(servidor__sistema_operativo__icontains=buscar)
        )
    
    filtro_versiones = Q()

    # FILTRO COMBINADO PARA BUSCAR VERSIONES DE PG Y SQL SERVER AL TIEMPO
    if version:
        filtro_versiones |= Q(major_version__icontains=version)

    if version_pg:
        filtro_versiones |= Q(major_version__icontains=version_pg)

    if filtro_versiones:
        all_inventario = all_inventario.filter(filtro_versiones)

    # FILTRO SI ESTA ADMINISTRADA LA INSTANCIA 
    if administrado in ["True", "False"]:
        all_inventario = all_inventario.filter(
            administrado=(administrado == "True")
        )

    # FILTRO POR CLIENTE
    if clientes_obj:
        all_inventario = all_inventario.filter(
            servidor__cliente__nombre__icontains=clientes_obj
        )

    # leer resultados de actualización si existen
    success_msgs = request.session.pop("actualizacion_success", [])
    error_msgs   = request.session.pop("actualizacion_errors", [])

    for msg in success_msgs:
        messages.success(request, msg)

    for msg in error_msgs:
        messages.error(request, msg)

    return render(request, 'paginas/inventario.html',{
        'instancias': all_inventario,
        'clientes': all_clientes,
        # devolver filtros al template
        'buscar': buscar,
        'versiones_sql': version,
        'versiones_pg': version_pg,
        'administrados': administrado,
        'cliente_actual': clientes_obj,
    })

#=================================================================================
#=================================================================================
#=================================================================================
#=================================================================================

def stream_actualizacion(request):
    log_queue    = queue.Queue()
    session_key  = request.session.session_key

    # asegurar que la sesión existe antes de pasarla al hilo
    if not session_key:
        request.session.create()
        session_key = request.session.session_key

    def run():
        errores_sync      = []
        errores_check     = []
        nuevos_servidores = 0
        nuevas_instancias = 0
        checks_realizados = 0

        try:
            result = actualizar_instancias_desde_conexiones(log_queue)

            nuevos_servidores = result["servidores_nuevos"]
            nuevas_instancias = result["instancias_nuevas"]
            errores_sync      = result["errores"]

            log_queue.put(f"✔ Servidores nuevos: {nuevos_servidores} | Instancias nuevas: {nuevas_instancias}")

            if not errores_sync:
                log_queue.put("✔ Sincronización completada sin errores.")
            else:
                for e in errores_sync:
                    log_queue.put(f"✘ {e}")

            log_queue.put("─── Iniciando verificación de builds ───")

            instancias_qs = instancia.objects.all()

            for inst in instancias_qs:
                try:
                    if inst.edition != 'PostgreSQL':
                        match = re.search(r'\d{4}', inst.major_version)
                        version_normalizada = match.group() if match else inst.major_version
                    else:
                        match = re.search(r'\d+', inst.major_version)
                        version_normalizada = match.group() if match else inst.major_version

                    ultima_actualizacion = actualizaciones.objects.filter(
                        major_version=version_normalizada,
                    ).order_by('-release_date', '-build').first()

                    if not ultima_actualizacion:
                        msg = f"Sin referencia para versión {inst.major_version} en {inst.nombre_instancia}"
                        errores_check.append(msg)
                        log_queue.put(f"✘ {msg}")
                        continue

                    build_instancia      = inst.build
                    build_referencia     = ultima_actualizacion.build
                    nombre_ultima_update = ultima_actualizacion.descripcion
                    fecha_publicacion    = ultima_actualizacion.release_date

                    if build_instancia == build_referencia:
                        estado  = "Actualizado"
                        detalle = f"La instancia está en el build más reciente ({build_referencia})"
                    else:
                        estado  = "Desactualizado"
                        detalle = f"Build actual: {build_instancia} | Build más reciente: {build_referencia} ({ultima_actualizacion.kb})"

                    checkSQL.objects.update_or_create(
                        instancia=inst,
                        defaults={
                            "build_detectado":      build_instancia,
                            "build_referencia":     build_referencia,
                            "estado":               estado,
                            "nombre_ultima_update": nombre_ultima_update,
                            "fecha_publicacion":    fecha_publicacion,
                            "fecha_check":          timezone.now(),
                            "detalle":              detalle
                        }
                    )

                    log_queue.put(f"✔ {inst.nombre_instancia} — {estado} (build {build_instancia})")
                    checks_realizados += 1

                except Exception as e:
                    msg = f"Error en {inst.nombre_instancia}: {str(e)}"
                    errores_check.append(msg)
                    log_queue.put(f"✘ {msg}")

            log_queue.put(f"✔ Checks realizados: {checks_realizados} | Errores: {len(errores_check)}")

        except Exception as e:
            log_queue.put(f"✘ Error inesperado: {str(e)}")
            errores_check.append(f"Error inesperado: {str(e)}")

        finally:
            # ── Guardar en sesión directamente desde el hilo ──────────────
            try:
                from importlib import import_module
                from django.conf import settings

                SessionStore = import_module(settings.SESSION_ENGINE).SessionStore
                session = SessionStore(session_key)

                success_msgs = [
                    f"Servidores nuevos: {nuevos_servidores} | Instancias nuevas: {nuevas_instancias}",
                    f"Checks realizados: {checks_realizados}",
                ]
                if not errores_sync and not errores_check:
                    success_msgs.append("Inventario actualizado sin errores.")

                session["actualizacion_success"] = success_msgs
                session["actualizacion_errors"]  = errores_sync + errores_check
                session.save()

            except Exception as e_session:
                print(f"Error guardando sesión: {e_session}")

            finally:
                log_queue.put(None)  # señal de fin

    thread = threading.Thread(target=run, daemon=True)
    thread.start()

    def event_stream():
        while True:
            mensaje = log_queue.get()
            if mensaje is None:
                yield "event: fin\ndata: Proceso completado\n\n"
                break
            yield f"data: {mensaje}\n\n"

    response = StreamingHttpResponse(event_stream(), content_type="text/event-stream")
    response["Cache-Control"] = "no-cache"
    response["X-Accel-Buffering"] = "no"
    return response
#=================================================================================
#=================================================================================
#=================================================================================
#=================================================================================
def detalles_instancia(request):

    if request.method == "POST":

        instancia_id = request.POST.get("id_instancia")
        servidor_id = request.POST.get("id_servidor")
        nombre_cliente = request.POST.get("cliente_id").strip()
        administrado = request.POST.get("administrado")
        observaciones = request.POST.get("observaciones").strip()

        if not nombre_cliente or not administrado:

            messages.error(
                request,
                "Debe llenar los campos obligatorios"
            )

            return redirect("listar_inventario")

        # obtener servidor
        servidor_obj = servidor.objects.get(id_servidor=servidor_id)

        # SI YA TIENE CLIENTE -> modificar nombre
        if nombre_cliente:
            servidor_obj.cliente_id = nombre_cliente
            servidor_obj.save()

        #obtener instancia
        instancia_obj = instancia.objects.get(id_instancia=instancia_id)
        instancia_obj.administrado = administrado
        instancia_obj.observaciones = observaciones
        instancia_obj.save()
        messages.success(
            request,
            "Informacion actualizadada correctamente."
        )


    return redirect("listar_inventario")

#=================================================================================
#=================================================================================
#=================================================================================
#=================================================================================

def eliminar_instancia(request):

    if request.method == 'POST':
        instancia_id = request.POST.get('id_instancia')
        instancia_obj= instancia.objects.get(id_instancia=instancia_id)
        
        if instancia_id and instancia_obj:
            instancia_obj.delete()
            messages.success(request,"sea eliminado la instancia correctamente")
            return redirect('listar_inventario')
        else:
            messages.error(request,"Error, al eliminar la instancia")
            return redirect('listar_inventario')

#=================================================================================
#=================================================================================
#=================================================================================
#=================================================================================

def exportar_excel(request):

    instancias_obj = instancia.objects.select_related(
        'servidor', 
        'servidor__cliente'
    ).prefetch_related(
        'servicios',
        'checksql'
    ).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Instancias de Bases de datos"

    encabezados = [
        "CLIENTE",
        "IP",
        "HOSTNAME",
        "INSTANCIA",
        "PUERTO",
        "S.O",
        "BUILD",
        "ULTIMA ACTUALIZACION",
        "CHECK",
        "EDICION",
        "ADMINISTRADO",
        "OBSERVACIONES",
        "SERVICIOS"
    ]

    ws.append(encabezados)

    # ====================================================
    #           ESTILOS
    # ====================================================

    encabezado_fill = PatternFill(
        fill_type="solid",
        start_color="1F4E78",
        end_color="1F4E78"
    )

    encabezado_font = Font(
        color="FFFFFF",
        bold=True
    )

    borde = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    # Alineación centrada (horizontal y vertical) + ajuste de texto
    alineacion_centrada = Alignment(
        horizontal="center",
        vertical="center",      # <-- CENTRADO VERTICAL (no bottom)
        wrap_text=True          # <-- AJUSTE DE TEXTO ACTIVADO
    )

    for celda in ws[1]:
        celda.fill = encabezado_fill
        celda.font = encabezado_font
        celda.border = borde
        celda.alignment = alineacion_centrada

    # ====================================================
    #           ESTILOS CONDICIONALES PARA CHECKS
    # ====================================================

    # Verde = actualizado
    fill_verde = PatternFill(
        fill_type="solid",
        start_color="C6EFCE",   # Verde claro de fondo
        end_color="C6EFCE"
    )
    font_verde = Font(
        color="006100",         # Verde oscuro de texto
        bold=True
    )

    # Rojo = desactualizado
    fill_rojo = PatternFill(
        fill_type="solid",
        start_color="FFC7CE",   # Rojo claro de fondo
        end_color="FFC7CE"
    )
    font_rojo = Font(
        color="9C0006",         # Rojo oscuro de texto
        bold=True
    )

    # ====================================================
    #           CALCULAR ANCHOS MÁXIMOS
    # ====================================================

    anchos = {}

    for idx, encabezado in enumerate(encabezados, 1):
        anchos[idx] = len(encabezado) + 2

    # ====================================================
    #           DATOS
    # ====================================================

    for inst in instancias_obj:

        check = inst.checksql.first()

        if check:
            checks = f"{check.nombre_ultima_update} | {check.build_referencia} | {check.fecha_publicacion}"
            estado = check.estado
        else:
            checks = ""
            estado = ""

        servicios = "\n".join(
            f"{s.nombre_servicio} | {s.estado_servicio} | {s.tipo_inicio}"
            for s in inst.servicios.all()
        )

        nombre_cliente = ""
        if inst.servidor.cliente:
            nombre_cliente = inst.servidor.cliente.nombre

        fila = [
            nombre_cliente,
            inst.servidor.ip,
            inst.servidor.hostname,
            inst.nombre_instancia,
            inst.puerto,
            inst.servidor.sistema_operativo,
            inst.build,
            checks,
            estado,
            inst.major_version,
            "SI" if inst.administrado else "NO",
            inst.observaciones,
            servicios
        ]

        ws.append(fila)

        # Obtener el número de la fila que acabamos de agregar
        numero_fila = ws.max_row

        # ====================================================
        #           APLICAR ESTILO CONDICIONAL
        # ====================================================

        celdas_a_pintar = [
            ws.cell(row=numero_fila, column=8),   # CHECKS
            ws.cell(row=numero_fila, column=9),   # ESTADO
        ]

        if estado and str(estado).lower() in ["actualizado", "ok", "updated", "up to date"]:
            for celda in celdas_a_pintar:
                celda.fill = fill_verde
                celda.font = font_verde

        elif estado and str(estado).lower() in ["desactualizado", "outdated", "pendiente", "old"]:
            for celda in celdas_a_pintar:
                celda.fill = fill_rojo
                celda.font = font_rojo

        # ====================================================
        #           APLICAR ESTILOS GENERALES A TODAS LAS CELDAS
        # ====================================================

        for col_idx in range(1, len(encabezados) + 1):
            celda = ws.cell(row=numero_fila, column=col_idx)

            # Aplicar borde y alineación a todas
            celda.border = borde
            celda.alignment = alineacion_centrada

            # Calcular ancho máximo
            if celda.value:
                lineas = str(celda.value).split('\n')
                max_linea = max(len(linea) for linea in lineas)
                if max_linea > anchos.get(col_idx, 0):
                    anchos[col_idx] = max_linea

    # ====================================================
    #           APLICAR ANCHOS DE COLUMNAS
    # ====================================================

    for idx, ancho in anchos.items():
        letra = get_column_letter(idx)
        ancho_final = min(ancho + 3, 60)
        ws.column_dimensions[letra].width = ancho_final

    # ====================================================
    #           AJUSTAR ALTURA DE FILAS
    # ====================================================

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        max_lineas = 1
        for celda in row:
            if celda.value and '\n' in str(celda.value):
                num_lineas = str(celda.value).count('\n') + 1
                if num_lineas > max_lineas:
                    max_lineas = num_lineas
        
        ws.row_dimensions[row[0].row].height = max(max_lineas * 15, 15)

    # ====================================================
    #           RESPUESTA
    # ====================================================

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="Inventario_BDS.xlsx"'
    wb.save(response)

    return response